"""XLSX cached cell-value extraction without formula evaluation."""

from __future__ import annotations

import re
from io import BytesIO
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils.exceptions import InvalidFileException
from lxml.etree import XMLSyntaxError

from ato_service.extraction.errors import ExtractionError
from ato_service.extraction.safety_xml import local_name, parse_xml_bounded, reject_external_relationship_targets
from ato_service.extraction.safety_zip import ZipMember, open_safe_zip, read_zip_member
from ato_service.extraction.types import ExtractedSegment, ExtractionLimits

_CELL_REF_PATTERN = re.compile(r"^([A-Z]{1,4})([1-9][0-9]{0,6})$")
_SHEET_PATH_PREFIX = "xl/worksheets/"


def extract_xlsx(content: bytes, *, limits: ExtractionLimits) -> list[ExtractedSegment]:
    """Extract cached cell values from one XLSX workbook."""
    members = open_safe_zip(content, limits=limits, office_container=True)
    for path in members:
        if path.endswith(".rels"):
            reject_external_relationship_targets(members[path].data)

    sheet_paths = _sorted_sheet_paths(members)
    if not sheet_paths:
        raise ExtractionError("xlsx contains no worksheets", error_code="source_parse_failed")

    formula_cells_by_path = _formula_cells_by_path(members, sheet_paths, limits=limits)
    workbook = _open_workbook(content)

    segments: list[ExtractedSegment] = []
    segment_index = 0
    try:
        worksheets = list(workbook.worksheets)
        if len(worksheets) != len(sheet_paths):
            raise ExtractionError("xlsx worksheet count mismatch", error_code="source_parse_failed")

        for worksheet in worksheets:
            sheet_path = _worksheet_path(worksheet)
            if sheet_path not in members:
                raise ExtractionError(
                    "xlsx worksheet path is invalid",
                    error_code="source_parse_failed",
                )
            sheet_name = _sheet_name_from_path(sheet_path)
            formula_cells = formula_cells_by_path.get(sheet_path, set())
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell, EmptyCell):
                        continue
                    ref = cell.coordinate
                    if not _CELL_REF_PATTERN.match(ref):
                        continue
                    value = _cached_cell_text(cell.value)
                    has_formula = ref in formula_cells
                    if value is None or value == "":
                        continue
                    segment_index += 1
                    segments.append(
                        ExtractedSegment(
                            segment_index=segment_index,
                            text=value,
                            locator={
                                "kind": "sheet_cell",
                                "sheet": sheet_name,
                                "cell": ref,
                            },
                            extraction_method="deterministic",
                            metadata=(
                                {"formula_ignored": True}
                                if has_formula
                                else None
                            ),
                        )
                    )
    except ExtractionError:
        raise
    except (KeyError, OSError, ParseError, TypeError, ValueError, XMLSyntaxError) as exc:
        raise ExtractionError("xlsx worksheet parsing failed", error_code="source_parse_failed") from exc
    finally:
        workbook.close()

    if not segments:
        raise ExtractionError("xlsx contains no cached cell values", error_code="source_parse_failed")
    return segments


def _sorted_sheet_paths(members: dict[str, ZipMember]) -> list[str]:
    return sorted(
        path
        for path in members
        if path.startswith(_SHEET_PATH_PREFIX) and path.endswith(".xml")
    )


def _sheet_name_from_path(path: str) -> str:
    basename = path.rsplit("/", 1)[-1]
    return basename.removesuffix(".xml")


def _worksheet_path(worksheet: object) -> str:
    path = getattr(worksheet, "_worksheet_path", None)
    if not isinstance(path, str) or not path:
        raise ExtractionError(
            "xlsx worksheet path is unavailable",
            error_code="source_parse_failed",
        )
    return path.lstrip("/")


def _formula_cells_by_path(
    members: dict[str, ZipMember],
    sheet_paths: list[str],
    *,
    limits: ExtractionLimits,
) -> dict[str, set[str]]:
    formula_cells_by_path: dict[str, set[str]] = {}
    for sheet_path in sheet_paths:
        root = parse_xml_bounded(read_zip_member(members, sheet_path), limits=limits)
        formula_cells: set[str] = set()
        for cell in root.iter():
            if local_name(cell.tag) != "c":
                continue
            ref = cell.attrib.get("r", "")
            if not _CELL_REF_PATTERN.match(ref):
                continue
            for child in list(cell):
                if local_name(child.tag) == "f":
                    formula_cells.add(ref)
                    break
        formula_cells_by_path[sheet_path] = formula_cells
    return formula_cells_by_path


def _open_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError, TypeError, ValueError) as exc:
        raise ExtractionError(
            "xlsx workbook is malformed",
            error_code="source_parse_failed",
        ) from exc


def _cached_cell_text(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text or None
